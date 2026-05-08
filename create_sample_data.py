"""Run once to create sample Excel files in the data/ directory."""

import os
import openpyxl

os.makedirs("data", exist_ok=True)

# ── products.xlsx ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["product", "id", "rule name"])
ws.append([
    "Aetna MAPD", 1,
    "agent gets 40%, his reporting manager gets 30%, his reporting manager's manager gets 20%, agency gets 10%",
])
ws.append([
    "Humana PPO", 2,
    "the agent receives 50%, the reporting manager receives 30%, the agency receives 20%",
])
ws.append([
    "UnitedHealth HMO", 3,
    "40% goes to the agent, 30% goes to the direct manager, 20% goes to the manager's manager, 10% goes to the agency",
])
ws.append([
    "Cigna PDP", 4,
    "agent gets 60%, agency gets 40%",
])
ws.append([
    "BlueCross Medigap", 5,
    "1) agent: 35%  2) manager: 35%  3) manager's manager: 20%  4) agency: 10% and subagent get 50%-50% of agents commission",
])
wb.save("data/products.xlsx")
print("Created data/products.xlsx")

# ── agents.xlsx ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["id", "name", "reporting id"])
ws.append([4, "Tom",   1])     # reports to Tony
ws.append([1, "Tony", 2])      # reports to thilak
ws.append([2, "Thilak",  3])   # reports to praveen
ws.append([3, "Praveen", "-"]) # top of chain
wb.save("data/agents.xlsx")
print("Created data/agents.xlsx (tom→tony→thilak→praveen)")

# ── statement.xlsx ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["product", "policy", "commission", "agent id"])

# Aetna MAPD  — 3-level rule (40/30/20/10)
ws.append(["Aetna MAPD",        "P-1001", 100, 1])   # full chain: TONY→THILAK→PRAVEEN
ws.append(["Aetna MAPD",        "P-1002", 200, 2])   # 2-agent chain: THILAK→PRAVEEN (rollup lvl 0)
ws.append(["Aetna MAPD",        "P-1003", 300, 3])   # 1-agent chain: PRAVEEN only  (rollup lvl 0+1)
ws.append(["Aetna MAPD",        "P-1012", 200, 4])   # subagent: TOM under TONY(0) → 50/50 split

# Humana PPO  — 2-level rule (50/30/20)
ws.append(["Humana PPO",        "P-1004", 150, 1])   # full chain
ws.append(["Humana PPO",        "P-1005", 250, 2])   # 2-agent chain (rollup lvl 0)

# UnitedHealth HMO — 3-level rule (40/30/20/10)
ws.append(["UnitedHealth HMO",  "P-1006", 500, 1])   # full chain
ws.append(["UnitedHealth HMO",  "P-1007", 400, 3])   # 1-agent chain (rollup lvl 0+1)

# Cigna PDP — agent+agency only (60/40)
ws.append(["Cigna PDP",         "P-1008", 120, 1])   # full chain — managers get 0%
ws.append(["Cigna PDP",         "P-1009",  80, 2])   # 2-agent chain — managers get 0%

# BlueCross Medigap — 3-level rule (35/35/20/10)
ws.append(["BlueCross Medigap", "P-1010", 600, 1])   # full chain
ws.append(["BlueCross Medigap", "P-1011", 450, 4])   # 2-agent chain (rollup lvl 0)

# Intentional mismatch — should be skipped
ws.append(["Delta Dental",      "P-9999", 999, 4])

wb.save("data/statement.xlsx")
print("Created data/statement.xlsx (12 rows, 1 intentional mismatch)")
print("\nAll sample data ready. Run:  .venv/bin/python -m commission_split.main")
